"""
Reports and Dashboard API Views
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db.models import Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from apps.documents.models import Crate, Document
from apps.requests.models import Request
from apps.auth.permissions import IsActiveUser
from apps.auth.models import Unit


def create_excel_response(workbook, filename):
    """Helper function to create an Excel file HTTP response"""
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def style_header_row(ws, num_columns):
    """Apply styling to the header row"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsActiveUser])
def stored_documents_report(request):
    """
    Get report of all storage requests with full history
    Filtered by unit_id query param or user's unit
    """
    # Get all storage requests
    storage_requests = Request.objects.filter(
        request_type='Storage'
    ).select_related(
        'crate', 'crate__storage', 'crate__department', 'crate__created_by', 'unit',
        'approved_by', 'allocated_by'
    ).prefetch_related('crate__documents').order_by('-request_date')

    # Filter by unit_id query param or user's unit
    unit_id = request.query_params.get('unit_id')
    if unit_id:
        storage_requests = storage_requests.filter(unit_id=unit_id)
    elif request.user.unit:
        storage_requests = storage_requests.filter(unit=request.user.unit)
    else:
        storage_requests = storage_requests.none()

    # Additional filters
    department_id = request.query_params.get('department_id')
    if department_id:
        storage_requests = storage_requests.filter(crate__department_id=department_id)

    status = request.query_params.get('status')
    if status:
        storage_requests = storage_requests.filter(status=status)

    from_date = request.query_params.get('from_date')
    if from_date:
        storage_requests = storage_requests.filter(request_date__date__gte=from_date)

    to_date = request.query_params.get('to_date')
    if to_date:
        storage_requests = storage_requests.filter(request_date__date__lte=to_date)

    # Check if export is requested
    export_format = request.query_params.get('export')

    # Serialize data
    data = []
    for req in storage_requests:
        storage_location = 'Not Allocated'
        if req.crate.storage:
            storage_location = req.crate.storage.get_full_location()

        data.append({
            'request_id': req.id,
            'crate_id': req.crate.id,
            'barcode': req.crate.barcode,
            'location': storage_location,
            'unit': req.unit.unit_code,
            'unit_name': req.unit.unit_name,
            'department': req.crate.department.department_name if req.crate.department else '',
            'request_date': req.request_date,
            'requested_by': req.crate.created_by.full_name if req.crate.created_by else None,
            'approval_date': req.approval_date,
            'approved_by': req.approved_by.full_name if req.approved_by else None,
            'allocation_date': req.allocation_date,
            'allocated_by': req.allocated_by.full_name if req.allocated_by else None,
            'status': req.status,
            'creation_date': req.crate.creation_date,
            'destruction_date': req.crate.destruction_date,
            'document_count': req.crate.documents.count(),
            'documents': [
                {
                    'id': doc.id,
                    'name': doc.document_name,
                    'number': doc.document_number,
                    'type': doc.document_type
                }
                for doc in req.crate.documents.all()
            ]
        })

    # Handle Excel export
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Storage Report'

        # Headers
        headers = ['Request ID', 'Crate ID', 'Barcode', 'Unit', 'Department', 'Status',
                   'Request Date', 'Approved By', 'Approval Date', 'Allocated By', 'Allocation Date',
                   'Location', 'Destruction Date', 'Document Count', 'Documents']
        ws.append(headers)
        style_header_row(ws, len(headers))

        # Data rows
        for item in data:
            docs_str = '; '.join([f"{d['name']} ({d['number']})" for d in item['documents']])
            ws.append([
                item['request_id'],
                item['crate_id'],
                item['barcode'],
                item['unit'],
                item['department'],
                item['status'],
                str(item['request_date']) if item['request_date'] else '',
                item['approved_by'] or '',
                str(item['approval_date']) if item['approval_date'] else '',
                item['allocated_by'] or '',
                str(item['allocation_date']) if item['allocation_date'] else '',
                item['location'],
                str(item['destruction_date']) if item['destruction_date'] else '',
                item['document_count'],
                docs_str
            ])

        auto_adjust_column_width(ws)
        return create_excel_response(wb, 'storage_report.xlsx')

    return Response({
        'count': len(data),
        'results': data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsActiveUser])
def withdrawn_documents_report(request):
    """
    Get report of all withdrawal requests with full history
    Filtered by unit_id query param or user's unit
    """
    # Get ALL withdrawal requests
    withdrawals = Request.objects.filter(
        request_type='Withdrawal'
    ).select_related(
        'crate', 'crate__department', 'crate__created_by', 'unit', 'withdrawn_by', 'issued_by', 'approved_by'
    ).prefetch_related('documents', 'crate__documents').order_by('-request_date')

    # Filter by unit_id query param or user's unit
    unit_id = request.query_params.get('unit_id')
    if unit_id:
        withdrawals = withdrawals.filter(unit_id=unit_id)
    elif request.user.unit:
        withdrawals = withdrawals.filter(unit=request.user.unit)
    else:
        withdrawals = withdrawals.none()

    # Additional filters
    department_id = request.query_params.get('department_id')
    if department_id:
        withdrawals = withdrawals.filter(crate__department_id=department_id)

    status = request.query_params.get('status')
    if status:
        withdrawals = withdrawals.filter(status=status)

    from_date = request.query_params.get('from_date')
    if from_date:
        withdrawals = withdrawals.filter(request_date__date__gte=from_date)

    to_date = request.query_params.get('to_date')
    if to_date:
        withdrawals = withdrawals.filter(request_date__date__lte=to_date)

    # Check if export is requested
    export_format = request.query_params.get('export')

    data = []
    for withdrawal in withdrawals:
        data.append({
            'request_id': withdrawal.id,
            'crate_id': withdrawal.crate.id,
            'barcode': withdrawal.crate.barcode,
            'unit': withdrawal.unit.unit_code,
            'unit_name': withdrawal.unit.unit_name,
            'withdrawn_by': withdrawal.withdrawn_by.full_name if withdrawal.withdrawn_by else None,
            'withdrawn_by_email': withdrawal.withdrawn_by.email if withdrawal.withdrawn_by else None,
            'approved_by': withdrawal.approved_by.full_name if withdrawal.approved_by else None,
            'issued_by': withdrawal.issued_by.full_name if withdrawal.issued_by else None,
            'request_date': withdrawal.request_date,
            'approval_date': withdrawal.approval_date,
            'issue_date': withdrawal.issue_date,
            'expected_return_date': withdrawal.expected_return_date,
            'return_date': withdrawal.return_date,
            'purpose': withdrawal.purpose,
            'full_withdrawal': withdrawal.full_withdrawal,
            'status': withdrawal.status,
            'is_overdue': withdrawal.is_overdue(),
            'document_count': withdrawal.crate.documents.count() if withdrawal.full_withdrawal else withdrawal.documents.count(),
            'documents': [
                {
                    'id': doc.id,
                    'name': doc.document_name,
                    'number': doc.document_number,
                    'type': doc.document_type
                }
                for doc in (withdrawal.crate.documents.all() if withdrawal.full_withdrawal else withdrawal.documents.all())
            ]
        })

    # Handle Excel export
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Withdrawal Report'

        # Headers
        headers = ['Request ID', 'Crate ID', 'Barcode', 'Unit', 'Withdrawn By', 'Email', 'Approved By',
                   'Issued By', 'Request Date', 'Approval Date', 'Issue Date', 'Expected Return',
                   'Return Date', 'Purpose', 'Status', 'Overdue', 'Document Count']
        ws.append(headers)
        style_header_row(ws, len(headers))

        # Data rows
        for item in data:
            ws.append([
                item['request_id'],
                item['crate_id'],
                item['barcode'],
                item['unit'],
                item['withdrawn_by'] or '',
                item['withdrawn_by_email'] or '',
                item['approved_by'] or '',
                item['issued_by'] or '',
                str(item['request_date']) if item['request_date'] else '',
                str(item['approval_date']) if item['approval_date'] else '',
                str(item['issue_date']) if item['issue_date'] else '',
                str(item['expected_return_date']) if item['expected_return_date'] else '',
                str(item['return_date']) if item['return_date'] else '',
                item['purpose'] or '',
                item['status'],
                'Yes' if item['is_overdue'] else 'No',
                item['document_count']
            ])

        auto_adjust_column_width(ws)
        return create_excel_response(wb, 'withdrawal_report.xlsx')

    return Response({
        'count': len(data),
        'results': data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsActiveUser])
def overdue_returns_report(request):
    """
    Get report of overdue withdrawal returns
    Filtered by unit_id query param or user's unit
    """
    # Get all overdue withdrawals
    overdue_withdrawals = Request.objects.filter(
        request_type='Withdrawal',
        status='Issued',
        expected_return_date__lt=timezone.now()
    ).select_related(
        'crate', 'unit', 'withdrawn_by', 'issued_by'
    ).order_by('expected_return_date')

    # Filter by unit_id query param or user's unit
    unit_id = request.query_params.get('unit_id')
    if unit_id:
        overdue_withdrawals = overdue_withdrawals.filter(unit_id=unit_id)
    elif request.user.unit:
        overdue_withdrawals = overdue_withdrawals.filter(unit=request.user.unit)
    else:
        overdue_withdrawals = overdue_withdrawals.none()

    # Check if export is requested
    export_format = request.query_params.get('export')

    data = []
    for withdrawal in overdue_withdrawals:
        days_overdue = (timezone.now().date() - withdrawal.expected_return_date.date()).days

        data.append({
            'request_id': withdrawal.id,
            'crate_id': withdrawal.crate.id,
            'barcode': withdrawal.crate.barcode,
            'unit': withdrawal.unit.unit_code,
            'withdrawn_by': withdrawal.withdrawn_by.full_name,
            'withdrawn_by_email': withdrawal.withdrawn_by.email,
            'issue_date': withdrawal.issue_date,
            'expected_return_date': withdrawal.expected_return_date,
            'days_overdue': days_overdue,
            'purpose': withdrawal.purpose
        })

    # Handle Excel export
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Overdue Returns'

        # Headers
        headers = ['Request ID', 'Crate ID', 'Barcode', 'Unit', 'Withdrawn By', 'Email',
                   'Issue Date', 'Expected Return Date', 'Days Overdue', 'Purpose']
        ws.append(headers)
        style_header_row(ws, len(headers))

        # Data rows
        for item in data:
            ws.append([
                item['request_id'],
                item['crate_id'],
                item['barcode'],
                item['unit'],
                item['withdrawn_by'],
                item['withdrawn_by_email'],
                str(item['issue_date']) if item['issue_date'] else '',
                str(item['expected_return_date']) if item['expected_return_date'] else '',
                item['days_overdue'],
                item['purpose'] or ''
            ])

        auto_adjust_column_width(ws)
        return create_excel_response(wb, 'overdue_returns_report.xlsx')

    return Response({
        'count': len(data),
        'results': data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsActiveUser])
def destruction_schedule_report(request):
    """
    Get report of all destruction requests with full history
    Filtered by unit_id query param or user's unit
    """
    # Get ALL destruction requests
    destruction_requests = Request.objects.filter(
        request_type='Destruction'
    ).select_related(
        'crate', 'crate__storage', 'crate__department', 'crate__created_by', 'unit',
        'approved_by', 'allocated_by'
    ).prefetch_related('crate__documents').order_by('-request_date')

    # Filter by unit_id query param or user's unit
    unit_id = request.query_params.get('unit_id')
    if unit_id:
        destruction_requests = destruction_requests.filter(unit_id=unit_id)
    elif request.user.unit:
        destruction_requests = destruction_requests.filter(unit=request.user.unit)
    else:
        destruction_requests = destruction_requests.none()

    # Additional filters
    department_id = request.query_params.get('department_id')
    if department_id:
        destruction_requests = destruction_requests.filter(crate__department_id=department_id)

    status = request.query_params.get('status')
    if status:
        destruction_requests = destruction_requests.filter(status=status)

    from_date = request.query_params.get('from_date')
    if from_date:
        destruction_requests = destruction_requests.filter(request_date__date__gte=from_date)

    to_date = request.query_params.get('to_date')
    if to_date:
        destruction_requests = destruction_requests.filter(request_date__date__lte=to_date)

    # Check if export is requested
    export_format = request.query_params.get('export')

    data = []
    for req in destruction_requests:
        storage_location = 'Not Allocated'
        if req.crate.storage:
            storage_location = req.crate.storage.get_full_location()

        data.append({
            'request_id': req.id,
            'crate_id': req.crate.id,
            'barcode': req.crate.barcode,
            'unit': req.unit.unit_code,
            'unit_name': req.unit.unit_name,
            'department': req.crate.department.department_name if req.crate.department else '',
            'request_date': req.request_date,
            'requested_by': req.crate.created_by.full_name if req.crate.created_by else None,
            'approval_date': req.approval_date,
            'approved_by': req.approved_by.full_name if req.approved_by else None,
            'status': req.status,
            'destruction_date': req.crate.destruction_date,
            'location': storage_location,
            'document_count': req.crate.documents.count(),
            'documents': [
                {
                    'id': doc.id,
                    'name': doc.document_name,
                    'number': doc.document_number,
                    'type': doc.document_type
                }
                for doc in req.crate.documents.all()
            ]
        })

    # Handle Excel export
    if export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Destruction Report'

        # Headers
        headers = ['Request ID', 'Crate ID', 'Barcode', 'Unit', 'Department', 'Status',
                   'Request Date', 'Approved By', 'Approval Date', 'Destruction Date',
                   'Location', 'Document Count', 'Documents']
        ws.append(headers)
        style_header_row(ws, len(headers))

        # Data rows
        for item in data:
            docs_str = '; '.join([f"{d['name']} ({d['number']})" for d in item['documents']])
            ws.append([
                item['request_id'],
                item['crate_id'],
                item['barcode'],
                item['unit'],
                item['department'],
                item['status'],
                str(item['request_date']) if item['request_date'] else '',
                item['approved_by'] or '',
                str(item['approval_date']) if item['approval_date'] else '',
                str(item['destruction_date']) if item['destruction_date'] else '',
                item['location'],
                item['document_count'],
                docs_str
            ])

        auto_adjust_column_width(ws)
        return create_excel_response(wb, 'destruction_report.xlsx')

    return Response({
        'count': len(data),
        'results': data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsActiveUser])
def dashboard_kpis(request):
    """
    Get dashboard KPIs and statistics
    Filtered by unit_id query param or user's unit
    """
    # Get unit for filtering (from query param or user's unit)
    unit_id = request.query_params.get('unit_id')
    if unit_id:
        try:
            user_unit = Unit.objects.get(id=unit_id)
        except Unit.DoesNotExist:
            user_unit = None
    else:
        user_unit = request.user.unit

    # Total stored crates
    crates_query = Crate.objects.filter(status='Active')
    if user_unit:
        crates_query = crates_query.filter(unit=user_unit)
    total_crates = crates_query.count()

    # Total documents (count documents in crates of user's unit)
    if user_unit:
        total_documents = Document.objects.filter(
            crate__unit=user_unit,
            crate__status='Active'
        ).distinct().count()
    else:
        total_documents = 0

    # Withdrawals in progress
    withdrawals_query = Request.objects.filter(
        request_type='Withdrawal',
        status='Issued'
    )
    if user_unit:
        withdrawals_query = withdrawals_query.filter(unit=user_unit)
    withdrawals_in_progress = withdrawals_query.count()

    # Overdue returns
    overdue_query = Request.objects.filter(
        request_type='Withdrawal',
        status='Issued',
        expected_return_date__lt=timezone.now()
    )
    if user_unit:
        overdue_query = overdue_query.filter(unit=user_unit)
    overdue_returns = overdue_query.count()

    # Pending approvals
    pending_query = Request.objects.filter(status='Pending')
    if user_unit:
        pending_query = pending_query.filter(unit=user_unit)
    pending_approvals = pending_query.count()

    # Storage requests this month
    from datetime import datetime
    current_month_start = datetime(timezone.now().year, timezone.now().month, 1)
    storage_requests_query = Request.objects.filter(
        request_type='Storage',
        request_date__gte=current_month_start
    )
    if user_unit:
        storage_requests_query = storage_requests_query.filter(unit=user_unit)
    storage_requests_this_month = storage_requests_query.count()

    # Upcoming destructions (next 30 days)
    from datetime import timedelta
    next_month = timezone.now().date() + timedelta(days=30)
    destruction_query = Crate.objects.filter(
        destruction_date__lte=next_month,
        destruction_date__gte=timezone.now().date(),
        status='Active'
    )
    if user_unit:
        destruction_query = destruction_query.filter(unit=user_unit)
    upcoming_destructions = destruction_query.count()

    # Request statistics by type
    request_stats_query = Request.objects
    if user_unit:
        request_stats_query = request_stats_query.filter(unit=user_unit)
    request_stats = request_stats_query.values('request_type', 'status').annotate(
        count=Count('id')
    )

    # Organize request stats
    stats_by_type = {}
    for stat in request_stats:
        req_type = stat['request_type']
        if req_type not in stats_by_type:
            stats_by_type[req_type] = {}
        stats_by_type[req_type][stat['status']] = stat['count']

    return Response({
        'total_stored_crates': total_crates,
        'total_documents': total_documents,
        'withdrawals_in_progress': withdrawals_in_progress,
        'overdue_returns': overdue_returns,
        'pending_approvals': pending_approvals,
        'storage_requests_this_month': storage_requests_this_month,
        'upcoming_destructions': upcoming_destructions,
        'request_statistics': stats_by_type
    })
